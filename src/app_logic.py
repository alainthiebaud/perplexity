
from pathlib import Path
import pandas as pd
from pandas import ExcelFile

def _last_row_sum(df: pd.DataFrame):
    if df is None or df.empty: return 0.0
    vals = pd.to_numeric(df.iloc[-1].drop(labels=["Période"], errors="ignore"), errors="coerce").fillna(0)
    return float(vals.sum())

def _kwh_totals(xl: ExcelFile, sheet: str):
    if sheet not in xl.sheet_names: return (0.0,0.0,0.0)
    df = xl.parse(sheet); last = df.iloc[-1] if not df.empty else None
    if last is None: return (0.0,0.0,0.0)
    hp = float(pd.to_numeric(last.get("HP_kwh",0), errors="coerce") or 0.0)
    hc = float(pd.to_numeric(last.get("HC_kwh",0), errors="coerce") or 0.0)
    sol= float(pd.to_numeric(last.get("Solaire_kwh",0), errors="coerce") or 0.0)
    return hp,hc,sol

def _extract_water(xl: ExcelFile):
    water = {"m3_ef":0.0,"m3_ec":0.0,"price_ef":0.0,"price_ec":0.0,"price_eu":0.0}
    if "Eau Froide" in xl.sheet_names:
        water["m3_ef"] = _last_row_sum(xl.parse("Eau Froide"))
    if "Eau Chaude" in xl.sheet_names:
        water["m3_ec"] = _last_row_sum(xl.parse("Eau Chaude"))
    if "Frais Eau" in xl.sheet_names:
        df = xl.parse("Frais Eau")
        if not df.empty:
            last = df.iloc[-1]
            if "Consommation (CHF)" in df.columns:
                water["price_ef"] = float(pd.to_numeric(last["Consommation (CHF)"], errors="coerce") or 0.0)
                water["price_ec"] = water["price_ef"]
            if "Taxe épuration (CHF)" in df.columns:
                water["price_eu"] = float(pd.to_numeric(last["Taxe épuration (CHF)"], errors="coerce") or 0.0)
    return water

def _derive_from_si(si_path: Path):
    water = {"price_ef":0.0,"price_ec":0.0,"price_eu":0.0}
    energy = {"price_hp":0.0,"price_hc":0.0,"price_solar":0.0,"price_global_simple":0.0}
    try:
        if not si_path or not Path(si_path).exists(): return water, energy
        xls = pd.ExcelFile(si_path)
        if "Données" not in xls.sheet_names: return water, energy
        df = xls.parse("Données")
        if "Période" in df.columns: df = df.sort_values("Période")
        r = df.iloc[-1]
        def get(c): return float(pd.to_numeric(r[c], errors="coerce") or 0.0) if c in df.columns else 0.0
        water["price_ef"] = get("Consommation Eau"); water["price_ec"] = water["price_ef"]; water["price_eu"] = get("Taxe épuration Eau")
        hp = get("HP énergie") + get("HP acheminement")
        hc = get("HC énergie") + get("HC acheminement")
        common = get("Swissgrid")+get("Taxe fédérale énergies renouvelables")+get("Taxe cantonale énergie")+get("Taxe communale usage du sol")+get("Taxe communale éclairage public")+get("Taxes communales environnementales")+get("TVA")
        energy["price_hp"] = hp + common
        energy["price_hc"] = hc + common
        energy["price_solar"] = get("Prix kWh Solaire") or energy["price_hc"]
        energy["price_global_simple"] = (energy["price_hp"] + energy["price_hc"])/2.0 if (energy["price_hp"] and energy["price_hc"]) else 0.0
    except Exception:
        pass
    return water, energy

def _price_pac_weighted(xl: ExcelFile, energy: dict) -> float:
    hp,hc,sol = _kwh_totals(xl,"PAC")
    denom = hp+hc+sol
    if denom<=0: return float(energy.get("price_global_simple",0) or 0.0)
    return (hp*energy["price_hp"] + hc*energy["price_hc"] + sol*energy["price_solar"]) / denom

def _price_global_weighted_real(xl: ExcelFile, energy: dict) -> float:
    hp1,hc1,sol1 = _kwh_totals(xl,"PAC")
    hp2,hc2,sol2 = _kwh_totals(xl,"Communs")
    hp=hp1+hp2; hc=hc1+hc2; sol=sol1+sol2
    denom = hp+hc+sol
    if denom<=0: return float(energy.get("price_global_simple",0) or 0.0)
    return (hp*energy["price_hp"] + hc*energy["price_hc"] + sol*energy["price_solar"]) / denom

def _price_ec_with_heating(xl: ExcelFile, price_ef: float, price_kwh: float) -> float:
    try:
        if "PAC" not in xl.sheet_names or "Eau Chaude" not in xl.sheet_names: return price_ef
        hp,hc,sol = _kwh_totals(xl,"PAC")
        pac_kwh = hp+hc+sol
        ch = _last_row_sum(xl.parse("Chauffage")) if "Chauffage" in xl.sheet_names else 0.0
        rf = _last_row_sum(xl.parse("Refroidissement")) if "Refroidissement" in xl.sheet_names else 0.0
        m3_ec = _last_row_sum(xl.parse("Eau Chaude"))
        if m3_ec<=0 or price_kwh<=0: return price_ef
        dhw_kwh = max(pac_kwh - ch - rf, 0.0)
        return float(price_ef) + (dhw_kwh * price_kwh) / m3_ec
    except Exception:
        return price_ef

class CalculationEngine:
    def __init__(self, config_mgr):
        self.config = config_mgr

    def _infer_tenants_from_charges(self, xls: ExcelFile):
        candidates = set()
        for sh in ("Eau Froide","Eau Chaude","Chauffage","Refroidissement"):
            if sh in xls.sheet_names:
                cols = list(xls.parse(sh).columns)
                for c in cols:
                    if c not in ("Période","PAC","Communs","Total","TOTAL","HP_kwh","HC_kwh","Solaire_kwh"):
                        candidates.add(c)
        return sorted([c for c in candidates if isinstance(c,str) and c.strip()])

    def preview_tables(self, charges_path: Path, tenants_path: Path):
        xls = ExcelFile(charges_path)
        previews = {}
        for sh in xls.sheet_names:
            df = xls.parse(sh)
            if sh in ("Eau Froide","Eau Chaude"):
                if "Communs" in df.columns:
                    df["Communs_calc"] = df["Communs"]
                else:
                    df["Communs_calc"] = 0.0
                if "Commentaires" not in df.columns:
                    df["Commentaires"] = ""
            previews[sh] = df
        tenants = None
        if tenants_path and Path(tenants_path).exists():
            try:
                tenants = pd.read_excel(tenants_path)
            except Exception:
                tenants = None
        if tenants is None or tenants.empty:
            names = self._infer_tenants_from_charges(xls)
            if names:
                tenants = pd.DataFrame({"Appartement": names})
        return previews, tenants

    def build_repartitions(self, charges_path: Path, tenants_df: pd.DataFrame):
        xl = ExcelFile(charges_path)
        water = _extract_water(xl)
        si_water, si_energy = _derive_from_si(Path(self.config.load_user_setting("si_path","")))
        for k in ["price_ef","price_ec","price_eu"]:
            if (water.get(k) or 0)==0 and (si_water.get(k) or 0)>0: water[k] = si_water[k]
        energy = si_energy

        energy["price_global_real"] = _price_global_weighted_real(xl, energy)
        energy["price_pac"] = _price_pac_weighted(xl, energy)
        water["price_ec"] = _price_ec_with_heating(xl, float(water.get("price_ef",0) or 0), float(energy.get("price_global_real",0) or 0))

        pac_hp,pac_hc,pac_sol = _kwh_totals(xl,"PAC")
        com_hp,com_hc,com_sol = _kwh_totals(xl,"Communs")

        details = [
            {"N°":1,"Clé":"Prix kWh (global simple)","Description":"(HP+HC)/2 (CHF/kWh) — tous frais inclus","Valeur": round(energy["price_global_simple"],4)},
            {"N°":2,"Clé":"Prix kWh (global réel)","Description":"Pondéré HP/HC/Solaire sur PAC+Communs (CHF/kWh)","Valeur": round(energy["price_global_real"],4)},
            {"N°":3,"Clé":"Prix kWh Chauff./Refroid.","Description":"Pondéré PAC (HP/HC/Solaire) (CHF/kWh)","Valeur": round(energy["price_pac"],4)},
            {"N°":4,"Clé":"Prix m3 EF","Description":"Eau froide (CHF/m³)","Valeur": round(water["price_ef"],2)},
            {"N°":5,"Clé":"Prix m3 EC","Description":"Eau chaude (CHF/m³) incluant chauffage ECS","Valeur": round(water["price_ec"],2)},
            {"N°":6,"Clé":"Prix m3 EU","Description":"Eaux usées (CHF/m³) — appliqué sur EC","Valeur": round(water["price_eu"],2)},
            {"N°":7,"Clé":"PAC kWh HP/HC/Sol","Description":"Mix PAC (kWh)","Valeur": f"{pac_hp:.2f}/{pac_hc:.2f}/{pac_sol:.2f}"},
            {"N°":8,"Clé":"Communs kWh HP/HC/Sol","Description":"Mix Communs (kWh)","Valeur": f"{com_hp:.2f}/{com_hc:.2f}/{com_sol:.2f}"},
        ]
        details_df = pd.DataFrame(details)

        rows = []
        rows.append({"Appartement":"PAC","Chauffage_kWh":pac_hp+pac_hc,"Refroid_kWh":0.0,"Eau_froide_m3":0.0,"Eau_chaude_m3":0.0,"Montant_eau":0.0,"Montant_energie": round((pac_hp*energy["price_hp"] + pac_hc*energy["price_hc"] + pac_sol*energy["price_solar"]),2)})
        rows.append({"Appartement":"COMMUNS","Chauffage_kWh":com_hp+com_hc,"Refroid_kWh":0.0,"Eau_froide_m3":0.0,"Eau_chaude_m3":0.0,"Montant_eau":0.0,"Montant_energie": round((com_hp*energy["price_hp"] + com_hc*energy["price_hc"] + com_sol*energy["price_solar"]),2)})
        if tenants_df is None or tenants_df.empty or "Appartement" not in tenants_df.columns:
            names = self._infer_tenants_from_charges(xl)
            tenants_df = pd.DataFrame({"Appartement": names}) if names else pd.DataFrame()

        if tenants_df is not None and not tenants_df.empty:
            df_ef = xl.parse("Eau Froide") if "Eau Froide" in xl.sheet_names else pd.DataFrame()
            df_ec = xl.parse("Eau Chaude") if "Eau Chaude" in xl.sheet_names else pd.DataFrame()
            df_ch = xl.parse("Chauffage") if "Chauffage" in xl.sheet_names else pd.DataFrame()
            df_rf = xl.parse("Refroidissement") if "Refroidissement" in xl.sheet_names else pd.DataFrame()
            last_ef = df_ef.iloc[-1] if not df_ef.empty else pd.Series(dtype=float)
            last_ec = df_ec.iloc[-1] if not df_ec.empty else pd.Series(dtype=float)
            last_ch = df_ch.iloc[-1] if not df_ch.empty else pd.Series(dtype=float)
            last_rf = df_rf.iloc[-1] if not df_rf.empty else pd.Series(dtype=float)
            for _, t in tenants_df.iterrows():
                apt = t.get("Appartement","")
                ef = float(pd.to_numeric(last_ef.get(apt,0), errors="coerce") or 0.0)
                ec = float(pd.to_numeric(last_ec.get(apt,0), errors="coerce") or 0.0)
                ch = float(pd.to_numeric(last_ch.get(apt,0), errors="coerce") or 0.0)
                rf = float(pd.to_numeric(last_rf.get(apt,0), errors="coerce") or 0.0)
                montant_eau = ef*water["price_ef"] + ec*water["price_ec"] + ec*water["price_eu"]
                montant_energie = ch*energy["price_pac"] + rf*energy["price_pac"]
                rows.append({"Appartement":apt,"Chauffage_kWh":ch,"Refroid_kWh":rf,"Eau_froide_m3":ef,"Eau_chaude_m3":ec,"Montant_eau":round(montant_eau,2),"Montant_energie":round(montant_energie,2)})
        rep_df = pd.DataFrame(rows)
        if not rep_df.empty:
            rep_df["Total_CHF"] = rep_df.get("Montant_eau",0).fillna(0) + rep_df.get("Montant_energie",0).fillna(0)

        explain = []
        explain.append("— Détails des calculs (à copier dans la facture) —")
        explain.append("1) Prix kWh (global simple) = (Prix_HP + Prix_HC) / 2 (CHF/kWh). (tous frais inclus)")
        explain.append("2) Prix kWh (global réel) = (HP_tot×Prix_HP + HC_tot×Prix_HC + Sol_tot×Prix_Solaire) / (HP_tot + HC_tot + Sol_tot) [PAC+Communs].")
        explain.append(f"   • Totaux PAC+Communs: HP={pac_hp+com_hp:.2f} kWh ; HC={pac_hc+com_hc:.2f} kWh ; Sol={pac_sol+com_sol:.2f} kWh.")
        explain.append("3) Prix kWh Chauffage/Refroid. (PAC) = (HP_PAC×Prix_HP + HC_PAC×Prix_HC + Sol_PAC×Prix_Solaire) / (HP_PAC + HC_PAC + Sol_PAC).")
        explain.append(f"   • Mix PAC: HP={pac_hp:.2f} ; HC={pac_hc:.2f} ; Sol={pac_sol:.2f}.")
        explain.append(f"   • Mix Communs: HP={com_hp:.2f} ; HC={com_hc:.2f} ; Sol={com_sol:.2f}.")
        explain.append("4) EF : Prix_EF (CHF/m³) — 'Frais Eau'. 5) EC : Prix_EC = Prix_EF + ((kWh_PAC_HP+HC+Sol − kWh_Chauffage − kWh_Refroid.) × Prix_KWh(global réel)) / m³_EC.")
        explain.append("6) EU : Prix_EU (CHF/m³) appliqué sur EC (m³).")
        explain.append("7) Montant_énergie(app) = kWh_Chauffage × Prix_KWh_PAC + kWh_Refroid. × Prix_KWh_PAC.")
        explain.append("8) Montant_eau(app) = m³_EF × Prix_EF + m³_EC × Prix_EC + m³_EC × Prix_EU.")
        explain.append("9) Total(app) = Montant_énergie + Montant_eau.")
        details_text = "\n".join(explain)

        return rep_df, details_df, details_text

    def export_log(self, out_dir: Path, period_label: str, rep_df: pd.DataFrame):
        out_dir = Path(out_dir); out_dir.mkdir(parents=True, exist_ok=True)
        import datetime, pandas as pd
        year = period_label.split("-")[0] if "-" in period_label else str(datetime.date.today().year)
        log_path = out_dir / f"Historique_factures_{year}.xlsx"
        log_df = pd.DataFrame()
        if log_path.exists():
            try: log_df = pd.read_excel(log_path)
            except Exception: log_df = pd.DataFrame()
        add = rep_df.copy(); add["Période"] = period_label
        log_df = pd.concat([log_df, add], ignore_index=True)
        with pd.ExcelWriter(log_path, engine="openpyxl") as wr:
            log_df.to_excel(wr, index=False, sheet_name="Log")
        return str(log_path)

    def export_invoices_xlsx(self, out_dir: Path, period_label: str, rep_df: pd.DataFrame, tenant_infos: pd.DataFrame):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except Exception as e:
            raise ImportError("openpyxl manquant. Installez-le avec: pip install openpyxl") from e
        out_base = Path(out_dir) / "Factures"; out_base.mkdir(parents=True, exist_ok=True)
        files = []
        if rep_df is None or rep_df.empty or not any(a not in ("PAC","COMMUNS","",None) for a in rep_df.get("Appartement", [])):
            raise RuntimeError("Aucun appartement à facturer : vérifiez le fichier Locataires ou les colonnes d'appartements (EF/EC/Chauffage/Refroidissement).")
        for _, row in rep_df.iterrows():
            apt = row.get("Appartement","")
            if apt in ["PAC","COMMUNS","TOTAL",""]:
                continue
            wb = Workbook(); ws = wb.active; ws.title = "Facture"
            bold = Font(bold=True)
            ws["A1"]="Edi et Alain Thiébaud"; ws["A1"].font=bold
            ws["A2"]="Rue d'Orbe 68, 1400 Yverdon-les-Bains"
            ws["A4"]="Facture d'électricité - Décompte groupé"; ws["A4"].font=bold
            ws["A6"]="Appartement:"; ws["B6"]=apt
            ws["A7"]="Période:"; ws["B7"]=period_label
            ws.append([]); ws.append(["Élément","Montant (CHF)"]); ws["A9"].font=bold; ws["B9"].font=bold
            me = float(row.get("Montant_energie",0) or 0); ma = float(row.get("Montant_eau",0) or 0); tt = me+ma
            ws.append(["TOTAL ÉNERGIE", me])
            ws.append(["TOTAL EAU", ma])
            ws.append(["TOTAL À PAYER", tt]); ws["A12"].font=bold; ws["B12"].font=bold
            ws2 = wb.create_sheet("Détails"); ws2["A1"]="Détail explicatif (voir application)."
            fp = out_base / f"Facture_{apt}_{period_label}.xlsx"
            wb.save(fp); files.append(fp)
        return str(out_base), [str(p) for p in files]

    def export_invoices_pdf(self, out_dir: Path, period_label: str, rep_df: pd.DataFrame, tenant_infos: pd.DataFrame, header_title: str = "Facture d'électricité - Décompte groupé", version_label: str = "Chat151"):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas
            from reportlab.lib.units import mm
            from reportlab.lib import colors
            from reportlab.platypus import Table, TableStyle
        except Exception as e:
            raise ImportError("ReportLab n'est pas installé. Installez-le avec: pip install reportlab") from e
        def _safe_filename(name: str) -> str:
            return "".join(c for c in str(name) if c.isalnum() or c in (" ","-","_",".")).strip().replace(" ","_")
        out_base = Path(out_dir) / "Factures_PDF"; out_base.mkdir(parents=True, exist_ok=True)
        files = []
        if rep_df is None or rep_df.empty or not any(a not in ("PAC","COMMUNS","",None) for a in rep_df.get("Appartement", [])):
            raise RuntimeError("Aucun appartement à facturer : vérifiez le fichier Locataires ou les colonnes d'appartements (EF/EC/Chauffage/Refroidissement).")
        for _, row in rep_df.iterrows():
            apt = row.get("Appartement","")
            if apt in ["PAC","COMMUNS","TOTAL",""]:
                continue
            me = float(row.get("Montant_energie",0) or 0); ma = float(row.get("Montant_eau",0) or 0); tt = me+ma
            fn = _safe_filename(f"Facture_{apt}_{period_label}.pdf"); fp = out_base / fn
            c = canvas.Canvas(str(fp), pagesize=A4); W,H = A4; margin = 18*mm
            y = H - margin
            c.setFont("Helvetica-Bold", 12); c.drawString(margin, y, "Edi et Alain Thiébaud"); y -= 14
            c.setFont("Helvetica", 10); c.drawString(margin, y, "Rue d'Orbe 68, 1400 Yverdon-les-Bains"); y -= 24
            c.setFont("Helvetica-Bold", 14); c.drawString(margin, y, header_title); y -= 10
            c.line(margin, y, W-margin, y); y -= 16
            c.setFont("Helvetica", 11); c.drawString(margin, y, f"Appartement : {apt}"); y -= 14
            c.drawString(margin, y, f"Période : {period_label}"); y -= 24
            data = [["Élément","Montant (CHF)"], ["TOTAL ÉNERGIE", f"{me:,.2f}".replace(",", " ").replace(".", ",")], ["TOTAL EAU", f"{ma:,.2f}".replace(",", " ").replace(".", ",")], ["TOTAL À PAYER", f"{tt:,.2f}".replace(",", " ").replace(".", ",")]]
            tbl = Table(data, colWidths=[100*mm, 50*mm])
            tbl.setStyle(TableStyle([("GRID",(0,0),(-1,-1),0.5,colors.black),("BACKGROUND",(0,0),(-1,0),colors.lightgrey),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("ALIGN",(1,1),(1,-1),"RIGHT"),("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold")]))
            _, th = tbl.wrapOn(c, W-2*margin, y-margin); tbl.drawOn(c, margin, y-th); y = y-th-20
            c.setFont("Helvetica-Oblique", 8); c.drawString(margin, margin, f"Facture générée automatiquement ({version_label}).")
            c.showPage()
            y = H - margin; c.setFont("Helvetica-Bold", 12); c.drawString(margin, y, "Détails des calculs"); y -= 12
            c.setFont("Helvetica", 9)
            lines = [
                "1) Prix kWh (global simple) = (Prix_HP + Prix_HC) / 2 (CHF/kWh).",
                "2) Prix kWh (global réel) = pondération HP/HC/Solaire sur PAC + Communs.",
                "3) Prix kWh Chauffage/Refroidissement = pondération PAC (HP/HC/Solaire).",
                "4) EF : Prix_EF (CHF/m³) — 'Frais Eau'.",
                "5) EC : Prix_EC = Prix_EF + ((kWh_PAC_HP+HC+Sol − kWh_Chauffage − kWh_Refroid.) × Prix_KWh(global réel)) / m³_EC.",
                "6) EU : Prix_EU (CHF/m³) appliqué sur EC.",
                "7) Montant_énergie(app) = kWh_Chauffage × Prix_KWh_PAC + kWh_Refroid. × Prix_KWh_PAC.",
                "8) Montant_eau(app) = m³_EF × Prix_EF + m³_EC × Prix_EC + m³_EC × Prix_EU.",
                "9) Total(app) = Montant_énergie + Montant_eau."
            ]
            for ln in lines:
                if y < margin + 60:
                    c.showPage(); y = H - margin; c.setFont("Helvetica", 9)
                c.drawString(margin, y, ln); y -= 12
            y_qr = margin + 80; c.setFont("Helvetica-Oblique", 9); c.drawString(margin, y_qr + 50, "Zone QR (placeholder)")
            from reportlab.lib.units import mm as _mm
            c.rect(margin, y_qr, 50*_mm, 50*_mm, stroke=1, fill=0)
            c.setFont("Helvetica-Oblique", 8); c.drawString(margin, margin, "Placez ici le QR-bill si nécessaire.")
            c.showPage(); c.save(); files.append(fp)
        return str(out_base), [str(p) for p in files]
